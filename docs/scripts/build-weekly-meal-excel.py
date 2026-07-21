# -*- coding: utf-8 -*-
"""Excel: 早餐=豆浆/水果/面包 + 午餐 + 下午茶水果；不含晚餐夜宵."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "md" / "_weekly_meals_dump.json"
OUT = ROOT / "md" / "\u4e00\u5468\u996e\u98df\u70ed\u91cf\u4e00\u89c8.xlsx"
OUT_ASCII = ROOT / "md" / "weekly-meal-calories.xlsx"

STYLE = {
    "早餐": ("1B6B4A", "E8F5EF"),
    "午餐": ("B45309", "FFF7ED"),
    "下午茶水果": ("BE185D", "FDF2F8"),
    "练前香蕉": ("0369A1", "E0F2FE"),
}

THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


def paint(ws, row, ncols, bg, font, align=CENTER):
    for col in range(1, ncols + 1):
        c = ws.cell(row, col)
        c.fill = fill(bg)
        c.font = font
        c.alignment = align
        c.border = THIN


def build_main(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "一周热量"
    days = data["days"]
    days_data = data["daysData"]
    ncols = 2 + len(days)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, "一人份 · 一周饮食热量一览（早餐＝豆浆/水果/面包 · 午餐 · 下午茶水果 · 练前香蕉）")
    t.font = Font(name="Microsoft YaHei", size=14, bold=True, color="111827")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub = ws.cell(
        2,
        1,
        "早餐拆三行；午餐列菜品；下午茶＝水果派对；练前香蕉＝每天下午训练前快碳（周四与派对各半根）。不含晚餐、夜宵。",
    )
    sub.font = Font(name="Microsoft YaHei", size=9, color="6B7280")
    sub.alignment = LEFT
    ws.row_dimensions[2].height = 36

    for col, h in enumerate(["大类", "项目"] + days, 1):
        c = ws.cell(3, col, h)
        c.fill = fill("111827")
        c.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        c.alignment = CENTER
        c.border = THIN

    accent, soft = STYLE["早餐"]
    # 早餐小计
    ws.cell(4, 1, "早餐")
    ws.cell(4, 2, "小计")
    for i, d in enumerate(days_data):
        cell = ws.cell(4, 3 + i, int(d["breakfast"]["total"]))
        cell.number_format = '0" 千卡"'
    paint(ws, 4, ncols, accent, Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF"))
    ws.cell(4, 2).alignment = LEFT

    breakfast_rows = [
        (5, "豆浆", "soy"),
        (6, "水果", "fruit"),
        (7, "面包", "bread"),
    ]
    for row, label, key in breakfast_rows:
        ws.cell(row, 1, "早餐")
        ws.cell(row, 2, label)
        for i, d in enumerate(days_data):
            part = d["breakfast"][key]
            text = f"{part.get('detail') or label}\n约 {int(part['kcal'])} 千卡"
            ws.cell(row, 3 + i, text)
        paint(ws, row, ncols, soft, Font(name="Microsoft YaHei", size=9, color="1F2937"))
        ws.cell(row, 2).alignment = LEFT
        ws.row_dimensions[row].height = 44

    ws.merge_cells(start_row=4, start_column=1, end_row=7, end_column=1)
    c = ws.cell(4, 1)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font = Font(name="Microsoft YaHei", size=13, bold=True, color="FFFFFF")
    c.fill = fill(accent)

    # 午餐（小计 + 套餐名 + 分菜明细）
    accent, soft = STYLE["午餐"]
    lunch_start = 9
    ws.cell(lunch_start, 1, "午餐")
    ws.cell(lunch_start, 2, "小计")
    for i, d in enumerate(days_data):
        cell = ws.cell(lunch_start, 3 + i, int(d["lunch"]["kcal"]))
        cell.number_format = '0" 千卡"'
    paint(ws, lunch_start, ncols, accent, Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF"))
    ws.cell(lunch_start, 2).alignment = LEFT

    ws.cell(lunch_start + 1, 1, "午餐")
    ws.cell(lunch_start + 1, 2, "当日套餐")
    for i, d in enumerate(days_data):
        ws.cell(lunch_start + 1, 3 + i, d["lunch"].get("title") or "")
    paint(ws, lunch_start + 1, ncols, soft, Font(name="Microsoft YaHei", size=9, color="1F2937"))
    ws.cell(lunch_start + 1, 2).alignment = LEFT
    ws.row_dimensions[lunch_start + 1].height = 32

    max_lunch_items = max((len(d["lunch"].get("items") or []) for d in days_data), default=0)
    for slot in range(max_lunch_items):
        row = lunch_start + 2 + slot
        ws.cell(row, 1, "午餐")
        ws.cell(row, 2, f"菜品 {slot + 1}")
        for i, d in enumerate(days_data):
            items = d["lunch"].get("items") or []
            if slot < len(items):
                it = items[slot]
                kcal = it.get("kcal")
                text = it.get("food") or ""
                if kcal is not None:
                    text = f"{text}\n约 {int(kcal)} 千卡"
                ws.cell(row, 3 + i, text)
            else:
                ws.cell(row, 3 + i, "—")
        paint(ws, row, ncols, soft, Font(name="Microsoft YaHei", size=8, color="1F2937"))
        ws.cell(row, 2).alignment = LEFT
        ws.row_dimensions[row].height = 40

    lunch_end = lunch_start + 1 + max_lunch_items
    ws.merge_cells(start_row=lunch_start, start_column=1, end_row=lunch_end, end_column=1)
    c = ws.cell(lunch_start, 1)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font = Font(name="Microsoft YaHei", size=13, bold=True, color="FFFFFF")
    c.fill = fill(accent)

    # 下午茶水果
    tea_start = lunch_end + 2
    accent, soft = STYLE["下午茶水果"]
    ws.cell(tea_start, 1, "下午茶水果")
    ws.cell(tea_start, 2, "参考热量")
    for i, d in enumerate(days_data):
        cell = ws.cell(tea_start, 3 + i, int(d["tea"]["kcal"]))
        cell.number_format = '0" 千卡"'
    paint(ws, tea_start, ncols, accent, Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF"))
    ws.cell(tea_start, 2).alignment = LEFT

    ws.cell(tea_start + 1, 1, "下午茶水果")
    ws.cell(tea_start + 1, 2, "当日水果")
    for i, d in enumerate(days_data):
        ws.cell(tea_start + 1, 3 + i, d["tea"].get("note") or d["tea"].get("title") or "")
    paint(ws, tea_start + 1, ncols, soft, Font(name="Microsoft YaHei", size=9, color="1F2937"))
    ws.cell(tea_start + 1, 2).alignment = LEFT
    ws.row_dimensions[tea_start + 1].height = 36
    ws.merge_cells(start_row=tea_start, start_column=1, end_row=tea_start + 1, end_column=1)
    c = ws.cell(tea_start, 1)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    c.fill = fill(accent)

    # 练前香蕉
    banana_start = tea_start + 3
    accent, soft = STYLE["练前香蕉"]
    ws.cell(banana_start, 1, "练前香蕉")
    ws.cell(banana_start, 2, "参考热量")
    for i, d in enumerate(days_data):
        cell = ws.cell(banana_start, 3 + i, int(d["preWorkout"]["kcal"]))
        cell.number_format = '0" 千卡"'
    paint(ws, banana_start, ncols, accent, Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF"))
    ws.cell(banana_start, 2).alignment = LEFT

    ws.cell(banana_start + 1, 1, "练前香蕉")
    ws.cell(banana_start + 1, 2, "分量")
    for i, d in enumerate(days_data):
        pw = d["preWorkout"]
        ws.cell(banana_start + 1, 3 + i, f"{pw.get('detail')}\n{pw.get('note')}")
    paint(ws, banana_start + 1, ncols, soft, Font(name="Microsoft YaHei", size=9, color="1F2937"))
    ws.cell(banana_start + 1, 2).alignment = LEFT
    ws.row_dimensions[banana_start + 1].height = 40
    ws.merge_cells(start_row=banana_start, start_column=1, end_row=banana_start + 1, end_column=1)
    c = ws.cell(banana_start, 1)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    c.fill = fill(accent)

    # 合计
    total_row = banana_start + 3
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 2, "含练前香蕉")
    for i, total in enumerate(data["dailyTotals"]):
        cell = ws.cell(total_row, 3 + i, int(total))
        cell.number_format = '0" 千卡"'
    paint(ws, total_row, ncols, "0F172A", Font(name="Microsoft YaHei", size=12, bold=True, color="FDE68A"))
    ws.row_dimensions[total_row].height = 28

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    for i in range(len(days)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 24
    ws.freeze_panes = "C4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_summary(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("每日汇总")
    days = data["days"]
    days_data = data["daysData"]
    headers = ["项目"] + days + ["周均"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(1, 1, "早餐细分 + 午餐 + 下午茶水果 + 练前香蕉（千卡）")
    t.font = Font(name="Microsoft YaHei", size=14, bold=True, color="111827")

    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.fill = fill("111827")
        c.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        c.alignment = CENTER
        c.border = THIN

    rows = [
        ("早餐·豆浆", "1B6B4A", "E8F5EF", [int(d["breakfast"]["soy"]["kcal"]) for d in days_data]),
        ("早餐·水果", "1B6B4A", "E8F5EF", [int(d["breakfast"]["fruit"]["kcal"]) for d in days_data]),
        ("早餐·面包", "1B6B4A", "E8F5EF", [int(d["breakfast"]["bread"]["kcal"]) for d in days_data]),
        ("早餐小计", "0F766E", "CCFBF1", [int(d["breakfast"]["total"]) for d in days_data]),
        ("午餐", "B45309", "FFF7ED", [int(d["lunch"]["kcal"]) for d in days_data]),
        ("下午茶水果", "BE185D", "FDF2F8", [int(d["tea"]["kcal"]) for d in days_data]),
        ("练前香蕉", "0369A1", "E0F2FE", [int(d["preWorkout"]["kcal"]) for d in days_data]),
        ("全日合计", "0F172A", "0F172A", [int(x) for x in data["dailyTotals"]]),
    ]

    for r_i, (label, accent, soft, vals) in enumerate(rows):
        row = 3 + r_i
        avg = round(sum(vals) / len(vals)) if vals else 0
        ws.cell(row, 1, label)
        for i, v in enumerate(vals):
            ws.cell(row, 2 + i, v)
        ws.cell(row, 2 + len(days), avg)
        dark = label in ("早餐小计", "全日合计")
        for col in range(1, len(headers) + 1):
            c = ws.cell(row, col)
            c.fill = fill(accent if col == 1 or dark else soft)
            c.font = Font(
                name="Microsoft YaHei",
                size=10,
                bold=True,
                color="FFFFFF" if (col == 1 or dark) else "1F2937",
            )
            c.alignment = CENTER
            c.border = THIN

    ws.column_dimensions["A"].width = 14
    for i in range(1, len(headers)):
        ws.column_dimensions[get_column_letter(1 + i)].width = 10

    note = ws.cell(
        12,
        1,
        "说明：练前香蕉每天训练前 30～60 分钟；周四水果派对半根+练前半根，全天仍约 1 根。不含晚餐、夜宵。",
    )
    ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=len(headers))
    note.font = Font(name="Microsoft YaHei", size=9, color="6B7280")


def main() -> None:
    data = json.loads(DATA.read_text(encoding="utf-8"))
    wb = Workbook()
    build_main(wb, data)
    build_summary(wb, data)
    try:
        wb.save(OUT)
        print("wrote", OUT)
    except PermissionError:
        print("Chinese xlsx locked, skipped:", OUT)
    wb.save(OUT_ASCII)
    print("wrote", OUT_ASCII)
    print("dailyTotals", data["dailyTotals"])
    print("Tue lunch", data["daysData"][1]["lunch"]["title"], data["daysData"][1]["lunch"]["kcal"])


if __name__ == "__main__":
    main()
